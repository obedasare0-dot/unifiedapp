from __future__ import annotations

import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from typing import List

from app.services.product_validator import ValidationResult


class ValidationReporter:
    """Generates Excel reports for validation results."""
    
    def __init__(self, validation_results: List[ValidationResult], summary: dict):
        self.results = validation_results
        self.summary = summary
    
    def generate_excel_report(self) -> bytes:
        """Generate a comprehensive Excel report with validation results.
        
        Creates a multi-sheet workbook:
        - Summary: Overview of all checks
        - Failed Checks: Detailed view of failures
        - All Checks: Complete list of all validation results
        
        Returns:
            Excel file bytes
        """
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create sheets
        self._create_summary_sheet(wb)
        self._create_failed_checks_sheet(wb)
        self._create_all_checks_sheet(wb)
        
        # Save to bytes
        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()
    
    def _create_summary_sheet(self, wb: Workbook):
        """Create summary overview sheet."""
        ws = wb.create_sheet("Summary", 0)
        
        # Title
        ws['A1'] = 'PSA Data Validation Report'
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:D1')
        ws.row_dimensions[1].height = 30
        
        # Timestamp
        ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws['A2'].font = Font(size=10, italic=True)
        ws.merge_cells('A2:D2')
        
        # Overall Status
        ws['A4'] = 'Overall Status:'
        ws['A4'].font = Font(bold=True, size=12)
        ws['B4'] = self.summary['overall_status']
        ws['B4'].font = Font(bold=True, size=12)
        
        if self.summary['overall_status'] == 'PASS':
            ws['B4'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            ws['B4'].font = Font(bold=True, size=12, color="006100")
        else:
            ws['B4'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            ws['B4'].font = Font(bold=True, size=12, color="9C0006")
        
        # Summary Statistics
        ws['A6'] = 'Check Statistics'
        ws['A6'].font = Font(bold=True, size=11)
        ws['A6'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        ws.merge_cells('A6:B6')
        
        stats = [
            ('Total Checks Run:', self.summary['total_checks']),
            ('Passed:', self.summary['passed']),
            ('Failed:', self.summary['failed']),
            ('Warnings:', self.summary['warnings']),
            ('Total Errors Found:', self.summary['total_errors']),
        ]
        
        row = 7
        for label, value in stats:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            
            # Color code failures
            if label == 'Failed:' and value > 0:
                ws[f'B{row}'].font = Font(bold=True, color="9C0006")
            elif label == 'Passed:' and value > 0:
                ws[f'B{row}'].font = Font(bold=True, color="006100")
            
            row += 1
        
        # Failed Checks Summary
        failed_checks = [r for r in self.results if r.status == 'FAIL']
        if failed_checks:
            ws[f'A{row+1}'] = 'Failed Checks Summary'
            ws[f'A{row+1}'].font = Font(bold=True, size=11, color="9C0006")
            ws[f'A{row+1}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            ws.merge_cells(f'A{row+1}:D{row+1}')
            
            row += 2
            ws[f'A{row}'] = 'Check Name'
            ws[f'B{row}'] = 'Error Count'
            ws[f'C{row}'] = 'Message'
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].font = Font(bold=True)
                ws[f'{col}{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            
            row += 1
            for check in failed_checks:
                ws[f'A{row}'] = check.check_name
                ws[f'B{row}'] = check.error_count
                ws[f'C{row}'] = check.message
                ws[f'B{row}'].font = Font(bold=True, color="9C0006")
                row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 15
    
    def _create_failed_checks_sheet(self, wb: Workbook):
        """Create detailed sheet for failed checks only."""
        ws = wb.create_sheet("Failed Checks")
        
        failed_checks = [r for r in self.results if r.status == 'FAIL']
        
        if not failed_checks:
            ws['A1'] = 'No Failed Checks'
            ws['A1'].font = Font(size=14, bold=True, color="006100")
            ws['A1'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            return
        
        # Header
        headers = ['Check Name', 'Status', 'Error Count', 'Message', 'Details']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data
        row_idx = 2
        for check in failed_checks:
            ws.cell(row=row_idx, column=1, value=check.check_name)
            ws.cell(row=row_idx, column=2, value=check.status)
            ws.cell(row=row_idx, column=3, value=check.error_count)
            ws.cell(row=row_idx, column=4, value=check.message)
            ws.cell(row=row_idx, column=5, value=check.details)
            
            # Highlight error count
            ws.cell(row=row_idx, column=3).font = Font(bold=True, color="9C0006")
            
            # Highlight pass count (green)
            ws.cell(row=row_idx, column=3).font = Font(bold=True, color="006100")
            
            # Highlight error count (red if > 0)
            error_cell = ws.cell(row=row_idx, column=4)
            if check.error_count > 0:
                error_cell.font = Font(bold=True, color="9C0006")
            
            # Wrap text for details
            ws.cell(row=row_idx, column=6).alignment = Alignment(wrap_text=True, vertical='top')
            
            row_idx += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 50
        ws.column_dimensions['F'].width = 60
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    def _create_all_checks_sheet(self, wb: Workbook):
        """Create sheet with all validation checks."""
        ws = wb.create_sheet("All Checks")
        
        # Header
        headers = ['Check Name', 'Status', 'Pass Count', 'Error Count', 'Message', 'Details']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data
        row_idx = 2
        for check in self.results:
            ws.cell(row=row_idx, column=1, value=check.check_name)
            ws.cell(row=row_idx, column=2, value=check.status)
            ws.cell(row=row_idx, column=3, value=check.pass_count)
            ws.cell(row=row_idx, column=4, value=check.error_count)
            ws.cell(row=row_idx, column=5, value=check.message)
            ws.cell(row=row_idx, column=6, value=check.details)
            
            # Color code status
            status_cell = ws.cell(row=row_idx, column=2)
            if check.status == 'PASS':
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                status_cell.font = Font(bold=True, color="006100")
            elif check.status == 'FAIL':
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                status_cell.font = Font(bold=True, color="9C0006")
            elif check.status == 'WARNING':
                status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                status_cell.font = Font(bold=True, color="9C6500")
            
            # Highlight pass count (green)
            ws.cell(row=row_idx, column=3).font = Font(bold=True, color="006100")
            
            # Highlight error count (red if > 0)
            error_cell = ws.cell(row=row_idx, column=4)
            if check.error_count > 0:
                error_cell.font = Font(bold=True, color="9C0006")
            
            # Wrap text for details
            ws.cell(row=row_idx, column=6).alignment = Alignment(wrap_text=True, vertical='top')
            
            row_idx += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 50
        ws.column_dimensions['F'].width = 60
        
        # Freeze header row
        ws.freeze_panes = 'A2'


def generate_validation_report(
    validation_results: List[ValidationResult],
    summary: dict
) -> bytes:
    """Convenience function to generate validation report.
    
    Args:
        validation_results: List of ValidationResult objects
        summary: Summary dictionary from validator
    
    Returns:
        Excel file bytes
    """
    reporter = ValidationReporter(validation_results, summary)
    return reporter.generate_excel_report()
