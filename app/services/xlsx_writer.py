from __future__ import annotations

import io
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import numbers, Font, PatternFill, Alignment


def write_proofing_xlsx_bytes(
    col_names: list[str],
    data_rows: list[list[Any]],
    sheet_name: str,
) -> bytes:
    """Write proofing table with column headers and data.
    
    Args:
        col_names: List of column names (headers)
        data_rows: List of data rows
        sheet_name: Sheet name for the workbook
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Write header row
    for c_idx, col_name in enumerate(col_names, start=1):
        ws.cell(row=1, column=c_idx, value=col_name)

    # Write data rows
    for r_idx, row in enumerate(data_rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            # Format UPC column (column A) as text to preserve leading zeros
            if c_idx == 1:
                cell.number_format = '@'  # '@' is the format code for text

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def create_standard_headers(num_fields: int) -> list[str]:
    """Create standardized column headers for Product records.
    Uses ONLY the 6 documented field names from Confluence documentation.
    All other fields use generic Field_N naming.
    
    Known fields (matching original PSA file structure):
    - Field 0: Field_0 (contains 'Product')
    - Field 1: UPC
    - Field 5: Width_Inches
    - Field 6: Height_Inches
    - Field 7: Depth_Inches
    - Field 8: Color
    - Field 237: Front_Overhang_Inches
    """
    # ONLY the 6 documented fields from Walmart Confluence
    known_fields = {
        1: 'UPC',
        5: 'Width_Inches',
        6: 'Height_Inches',
        7: 'Depth_Inches',
        8: 'Color',
        237: 'Front_Overhang_Inches'
    }
    
    headers = []
    for i in range(num_fields):
        if i in known_fields:
            headers.append(known_fields[i])
        else:
            headers.append(f'Field_{i}')
    
    return headers


def write_raw_products_xlsx_bytes(product_rows: list[list[str]]) -> bytes:
    """Write raw product rows (all PSA fields) with standardized headers and professional styling."""
    if not product_rows:
        raise ValueError("No product rows")

    max_cols = max(len(r) for r in product_rows)
    
    # Use standardized headers
    headers = create_standard_headers(max_cols)

    # Create workbook with styling
    wb = Workbook()
    ws = wb.active
    ws.title = "Product Data"
    
    # Header styling - professional blue headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Write headers with styling
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Write data rows
    for row_idx, row in enumerate(product_rows, start=2):
        # Pad row if needed
        padded_row = row + [None] * (max_cols - len(row))
        for col_idx, val in enumerate(padded_row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            # Format UPC column (column B - Field 1) as text to preserve leading zeros
            if col_idx == 2:
                cell.number_format = '@'  # '@' is the format code for text
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Auto-adjust column widths for key columns
    column_widths = {
        'A': 15,  # Field_0 (Product)
        'B': 18,  # UPC
        'C': 15,  # Field_2
        'D': 30,  # Field_3 (usually description)
        'E': 15,  # Field_4
        'F': 15,  # Width_Inches
        'G': 15,  # Height_Inches
        'H': 15,  # Depth_Inches
        'I': 15,  # Color
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Save to bytes
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
