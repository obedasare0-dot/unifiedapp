"""Generate Excel files and ZIP for PSA data export."""
from __future__ import annotations

import io
import zipfile
from typing import Dict, Any

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment

from app.services.validation_reporter import generate_validation_report


def create_excel_export(data: Dict[str, Any]) -> bytes:
    """Create ZIP file with multi-tab PSA_Data.xlsx and combined validation report.
    
    Args:
        data: Dict from psa_processor containing:
            - product_df, product_validation, product_summary
            - planogram_df, planogram_validation, planogram_summary
            - fixture_df
            
    Returns:
        ZIP file bytes containing:
        - PSA_Data.xlsx (multi-tab: Product + Planogram + Fixture sheets)
        - Validation_Report.xlsx (combined checks from both tables)
    """
    
    print("[EXPORTER] Creating multi-tab Excel file...")
    
    excel_files = {}
    
    # Generate PSA_Data.xlsx with multiple sheets
    psa_data_out = io.BytesIO()
    with pd.ExcelWriter(psa_data_out, engine='openpyxl') as writer:
        
        # Sheet 1: Product Data
        if data['product_df'] is not None:
            data['product_df'].to_excel(writer, sheet_name='Product', index=False)
            ws_product = writer.sheets['Product']
            _style_header(ws_product)
            print(f"[EXPORTER] Added Product sheet ({len(data['product_df'])} rows, {len(data['product_df'].columns)} columns)")
        
        # Sheet 2: Planogram Data
        if data['planogram_df'] is not None:
            data['planogram_df'].to_excel(writer, sheet_name='Planogram', index=False)
            ws_planogram = writer.sheets['Planogram']
            _style_header(ws_planogram)
            
            # Highlight smart-mapped columns (7-10) in yellow
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            for col_idx in [8, 9, 10, 11]:  # Columns H, I, J, K
                ws_planogram.cell(row=1, column=col_idx).fill = yellow_fill
                ws_planogram.cell(row=1, column=col_idx).font = Font(bold=True, color="000000")
            
            print(f"[EXPORTER] Added Planogram sheet ({len(data['planogram_df'])} rows, {len(data['planogram_df'].columns)} columns)")
        
        # Sheet 3: Fixture Data
        if data.get('fixture_df') is not None:
            data['fixture_df'].to_excel(writer, sheet_name='Fixture', index=False)
            ws_fixture = writer.sheets['Fixture']
            _style_header(ws_fixture)
            print(f"[EXPORTER] Added Fixture sheet ({len(data['fixture_df'])} rows, {len(data['fixture_df'].columns)} columns)")
    
    excel_files['PSA_Data.xlsx'] = psa_data_out.getvalue()
    sheet_count = sum([1 for key in ['product_df', 'planogram_df', 'fixture_df'] if data.get(key) is not None])
    print(f"[EXPORTER] Created PSA_Data.xlsx with {sheet_count} sheets ({len(excel_files['PSA_Data.xlsx'])} bytes)")
    
    # Combine validation results
    all_validation_results = []
    combined_summary = {
        'total_checks': 0,
        'passed': 0,
        'failed': 0,
        'warnings': 0,
        'overall_status': 'PASS',  # Will update if any failures
        'total_errors': 0,  # Track total error count across all failed checks
        'total_records': 0  # Track total records processed
    }
    
    # Track total records
    if data['product_df'] is not None:
        combined_summary['total_records'] += len(data['product_df'])
    if data['planogram_df'] is not None:
        combined_summary['total_records'] += len(data['planogram_df'])
    if data.get('fixture_df') is not None:
        combined_summary['total_records'] += len(data['fixture_df'])
    
    # Add Product validation
    if data['product_validation']:
        # Prefix each check with table name
        for result in data['product_validation']:
            result.check_name = f"[Product] {result.check_name}"
            # Sum up error counts
            combined_summary['total_errors'] += result.error_count
        all_validation_results.extend(data['product_validation'])
        combined_summary['total_checks'] += data['product_summary']['total_checks']
        combined_summary['passed'] += data['product_summary']['passed']
        combined_summary['failed'] += data['product_summary']['failed']
        combined_summary['warnings'] += data['product_summary']['warnings']
    
    # Add Planogram validation
    if data['planogram_validation']:
        # Prefix each check with table name
        for result in data['planogram_validation']:
            result.check_name = f"[Planogram] {result.check_name}"
            # Sum up error counts
            combined_summary['total_errors'] += result.error_count
        all_validation_results.extend(data['planogram_validation'])
        combined_summary['total_checks'] += data['planogram_summary']['total_checks']
        combined_summary['passed'] += data['planogram_summary']['passed']
        combined_summary['failed'] += data['planogram_summary']['failed']
        combined_summary['warnings'] += data['planogram_summary']['warnings']
    
    # Add Fixture validation
    if data.get('fixture_validation'):
        # Prefix each check with table name
        for result in data['fixture_validation']:
            result.check_name = f"[Fixture] {result.check_name}"
            # Sum up error counts
            combined_summary['total_errors'] += result.error_count
        all_validation_results.extend(data['fixture_validation'])
        combined_summary['total_checks'] += data['fixture_summary']['total_checks']
        combined_summary['passed'] += data['fixture_summary']['passed']
        combined_summary['failed'] += data['fixture_summary']['failed']
        combined_summary['warnings'] += data['fixture_summary']['warnings']
    
    # Update overall status based on failures
    if combined_summary['failed'] > 0:
        combined_summary['overall_status'] = 'FAIL'
    elif combined_summary['warnings'] > 0:
        combined_summary['overall_status'] = 'WARNING'
    else:
        combined_summary['overall_status'] = 'PASS'
    
    # Generate combined validation report
    validation_report_bytes = generate_validation_report(all_validation_results, combined_summary)
    excel_files['Validation_Report.xlsx'] = validation_report_bytes
    print(f"[EXPORTER] Created Validation_Report.xlsx with {combined_summary['total_checks']} total checks")
    
    # Create ZIP
    zip_out = io.BytesIO()
    with zipfile.ZipFile(zip_out, mode='w', compression=zipfile.ZIP_DEFLATED) as zf:
        for filename, file_bytes in excel_files.items():
            zf.writestr(filename, file_bytes)
    
    zip_bytes = zip_out.getvalue()
    print(f"[EXPORTER] Generated ZIP ({len(zip_bytes)} bytes) with 2 files: PSA_Data.xlsx + Validation_Report.xlsx")
    
    return zip_bytes


def _style_header(ws):
    """Apply Walmart blue styling to header row."""
    header_fill = PatternFill(start_color="0053E2", end_color="0053E2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
